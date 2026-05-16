"""Ingest CIQUAL 2025 xlsx → compact JSON lookup.

Output: corpus/ciqual_2025.json  (full, lookup by alim_code)
         corpus/ciqual_2025_index.json  (light index for name search)

License: Etalab 2.0 / Anses 2025.
Source: https://doi.org/10.57745/RDMHWY

Usage: python3 ingest_ciqual.py
"""

from __future__ import annotations

import json
import re
import sys
from pathlib import Path
from openpyxl import load_workbook

SCRIPT_DIR = Path(__file__).resolve().parent
CORPUS_DIR = SCRIPT_DIR.parent / "corpus"
SOURCE = CORPUS_DIR / "ciqual_2025.xlsx"

# Headers in CIQUAL 2025 have embedded newlines and unit suffixes;
# we map them to clean JSON keys with explicit units.
# Each entry: (column_index, json_key, unit).
COLUMNS = [
    (0, "group_code", None),
    (1, "subgroup_code", None),
    (2, "subsubgroup_code", None),
    (3, "group_fr", None),
    (4, "subgroup_fr", None),
    (5, "subsubgroup_fr", None),
    (6, "code", None),
    (7, "name_fr", None),
    (8, "name_sci", None),
]

# Nutrient columns we keep (subset of the 74). All values are per 100 g.
NUTRIENT_COLS = [
    (10, "energy_kcal", "kcal"),       # Energie kcal règlement UE
    (13, "water_g", "g"),
    (14, "protein_g", "g"),
    (16, "carb_g", "g"),
    (17, "fat_g", "g"),
    (18, "sugar_g", "g"),
    (25, "starch_g", "g"),
    (26, "fiber_g", "g"),
    (29, "alcohol_g", "g"),
    (31, "saturated_fat_g", "g"),
    (32, "monounsat_fat_g", "g"),
    (33, "polyunsat_fat_g", "g"),
    (48, "cholesterol_mg", "mg"),
    (49, "salt_g", "g"),               # Sel chlorure de sodium (NaCl)
    (50, "calcium_mg", "mg"),
    (53, "iron_mg", "mg"),
    (54, "iodine_ug", "µg"),
    (55, "magnesium_mg", "mg"),
    (57, "phosphorus_mg", "mg"),
    (58, "potassium_mg", "mg"),
    (60, "sodium_mg", "mg"),
    (61, "zinc_mg", "mg"),
    (62, "vit_a_retinol_eq_ug", "µg"),
    (65, "vit_d_ug", "µg"),
    (68, "alpha_tocopherol_mg", "mg"),
    (72, "vit_c_mg", "mg"),
    (73, "vit_b1_mg", "mg"),
    (74, "vit_b2_mg", "mg"),
    (75, "vit_b3_mg", "mg"),
    (77, "vit_b6_mg", "mg"),
    (78, "vit_b9_dfe_ug", "µg"),       # Folates équivalents DFE
    (82, "vit_b12_ug", "µg"),
]


def parse_value(raw):
    """CIQUAL uses '<0.5', 'traces', '-' for special values. Return float or None."""
    if raw is None or raw == "":
        return None
    if isinstance(raw, (int, float)):
        return float(raw)
    s = str(raw).strip().replace(",", ".")
    if s in {"-", "traces", "0", ""}:
        return 0.0 if s in {"0", "traces"} else None
    m = re.match(r"^<\s*([\d.]+)$", s)
    if m:
        # "<0.5" → half of threshold (conservative-ish)
        return float(m.group(1)) / 2.0
    try:
        return float(s)
    except ValueError:
        return None


def main() -> None:
    if not SOURCE.exists():
        sys.exit(f"Source not found: {SOURCE}")
    wb = load_workbook(SOURCE, read_only=True, data_only=True)
    ws = wb["composition nutritionnelle"]
    rows = ws.iter_rows(values_only=True)
    next(rows)  # skip header

    full = {}
    index = []
    for row in rows:
        code = row[6]
        if code is None:
            continue
        code_str = str(code).strip()
        name_fr = (row[7] or "").strip()
        if not name_fr:
            continue
        nutrients = {}
        for col_idx, key, _unit in NUTRIENT_COLS:
            nutrients[key] = parse_value(row[col_idx])
        entry = {
            "code": code_str,
            "name_fr": name_fr,
            "name_sci": (row[8] or "").strip() or None,
            "group_fr": (row[3] or "").strip() or None,
            "subgroup_fr": (row[4] or "").strip() or None,
            "subsubgroup_fr": (row[5] or "").strip() or None,
            "nutrients_per_100g": nutrients,
        }
        full[code_str] = entry
        index.append({"code": code_str, "name_fr": name_fr})

    units = {key: unit for _, key, unit in NUTRIENT_COLS}
    units["energy_kcal"] = "kcal"

    out = {
        "_meta": {
            "source": "ANSES Ciqual 2025",
            "version_date": "2025-11-03",
            "doi": "10.57745/RDMHWY",
            "license": "Etalab 2.0",
            "citation": "Anses. 2025. Table de composition nutritionnelle des aliments Ciqual 2025. https://doi.org/10.57745/RDMHWY",
            "units": units,
            "count": len(full),
        },
        "foods": full,
    }

    out_full = CORPUS_DIR / "ciqual_2025.json"
    out_index = CORPUS_DIR / "ciqual_2025_index.json"

    out_full.write_text(json.dumps(out, ensure_ascii=False, separators=(",", ":")), encoding="utf-8")
    out_index.write_text(json.dumps({"_meta": out["_meta"], "items": index}, ensure_ascii=False, separators=(",", ":")), encoding="utf-8")

    print(f"Wrote {len(full)} foods to {out_full} ({out_full.stat().st_size / 1024:.0f} KB)")
    print(f"Wrote index to {out_index} ({out_index.stat().st_size / 1024:.0f} KB)")


if __name__ == "__main__":
    main()
